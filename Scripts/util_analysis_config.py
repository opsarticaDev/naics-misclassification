"""
OSHA NAICS Validation: Shared Analysis Utilities
===================================================
Common functions and data shared across steps 15, 16, and 17:
  - DART rate loading and cascading lookup
  - NAICS description tree building
  - Text tokenization with stemming and synonym expansion
  - Sector equivalence helpers

Import from this module to avoid code duplication.
"""

import csv
import os
import re
from collections import Counter

from util_pipeline_config import DATASET_YEAR

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
NAICS_2017_FILE = os.path.join(BASE_DIR, "Reference", "naics_2017_official.xlsx")
NAICS_2022_FILE = os.path.join(BASE_DIR, "Reference", "naics_2022_official.xlsx")
BLS_DART_FILE = os.path.join(BASE_DIR, "Reference", f"bls_dart_rates_{DATASET_YEAR}.xlsx")

# ---------------------------------------------------------------------------
# Stopwords for description token matching
# ---------------------------------------------------------------------------
ANALYSIS_STOPWORDS = {
    "the", "and", "for", "with", "all", "other", "not", "than",
    "that", "this", "are", "was", "were", "been", "being", "have",
    "has", "had", "does", "did", "but", "from", "into", "out",
    "own", "its", "etc", "including", "except", "related",
    "type", "types", "kind", "kinds", "made", "used", "using",
    "such", "also", "see", "like",
}

# ---------------------------------------------------------------------------
# Domain synonym bridge: maps terms that stemming can't connect.
# Used by tokenize() in steps 16 and 17 to bridge vocabulary gaps
# between establishment descriptions and official NAICS tree names.
# ---------------------------------------------------------------------------
SYNONYM_BRIDGE = {
    "airport": {"air", "transportation", "aviation"},
    "aviation": {"air", "transportation", "airport"},
    "dental": {"dentist", "dentistry", "oral"},
    "dentist": {"dental", "dentistry", "oral"},
    "pharmacy": {"pharmaceutical", "drug", "drugstore"},
    "pharmaceutical": {"pharmacy", "drug"},
    "trucking": {"freight", "truck", "hauling", "motor", "carrier"},
    "freight": {"trucking", "cargo", "shipping"},
    "railroad": {"rail", "railway", "locomotive"},
    "railway": {"rail", "railroad"},
    "plumbing": {"plumber", "piping", "heating", "cooling"},
    "hvac": {"heating", "cooling", "ventilation", "air", "conditioning"},
    "roofing": {"roofer", "roof"},
    "electric": {"electrical", "electricity", "power"},
    "electrical": {"electric", "electricity", "wiring"},
    "lumber": {"timber", "wood", "sawmill"},
    "timber": {"lumber", "wood", "logging"},
    "auto": {"automobile", "automotive", "vehicle", "car"},
    "automotive": {"auto", "automobile", "vehicle", "car"},
    "veterinary": {"animal", "pet", "vet"},
    "daycare": {"childcare", "child", "care", "preschool"},
    "childcare": {"daycare", "child", "care", "preschool"},
    "warehouse": {"warehousing", "storage", "distribution"},
    "warehousing": {"warehouse", "storage", "distribution"},
    "landscaping": {"landscape", "lawn", "garden"},
    "landscape": {"landscaping", "lawn", "garden"},
    "concrete": {"cement", "masonry"},
    "masonry": {"concrete", "cement", "brick", "stone"},
    "demolition": {"wrecking", "deconstruction"},
    "excavating": {"excavation", "earthwork", "grading"},
    "paving": {"asphalt", "pavement"},
    "hotel": {"motel", "lodging", "hospitality", "accommodation"},
    "motel": {"hotel", "lodging", "hospitality"},
    "restaurant": {"dining", "food", "eatery"},
    "cafeteria": {"food", "dining", "catering"},
    "catering": {"food", "cafeteria"},
    "hospital": {"medical", "surgical", "healthcare", "clinical"},
    "clinic": {"clinical", "medical", "healthcare", "outpatient"},
    "nursing": {"nurse", "geriatric", "elder", "care"},
    "ambulance": {"emergency", "paramedic", "ems"},
    "refinery": {"refining", "petroleum", "petrochemical"},
    "brewery": {"brewing", "beer", "craft"},
    "winery": {"wine", "vineyard", "viticulture"},
    "quarry": {"quarrying", "crushed", "stone", "mining"},
    "mining": {"mine", "mineral", "extraction"},
}

# ---------------------------------------------------------------------------
# Sector equivalence groups (combined NAICS sectors)
# ---------------------------------------------------------------------------
SECTOR_EQUIV = {
    "31": {"31", "32", "33"}, "32": {"31", "32", "33"}, "33": {"31", "32", "33"},
    "44": {"44", "45"}, "45": {"44", "45"},
    "48": {"48", "49"}, "49": {"48", "49"},
}

SECTOR_NAMES = {
    "11": "Agriculture", "21": "Mining", "22": "Utilities",
    "23": "Construction", "31": "Manufacturing", "32": "Manufacturing",
    "33": "Manufacturing", "42": "Wholesale Trade",
    "44": "Retail Trade", "45": "Retail Trade",
    "48": "Transportation", "49": "Transportation",
    "51": "Information", "52": "Finance/Insurance",
    "53": "Real Estate", "54": "Professional Services",
    "55": "Management", "56": "Admin/Support/Waste",
    "61": "Education", "62": "Healthcare/Social",
    "71": "Arts/Entertainment", "72": "Accommodation/Food",
    "81": "Other Services", "92": "Public Administration",
}


# ---------------------------------------------------------------------------
# Sector helpers
# ---------------------------------------------------------------------------

def get_sector(naics_code):
    """Extract the 2-digit NAICS sector."""
    return str(naics_code).strip()[:2]


def sectors_match(s1, s2):
    """Check if two 2-digit sectors are equivalent (e.g., 31==32==33)."""
    equiv = SECTOR_EQUIV.get(s1, {s1})
    return s2 in equiv


def sector_name(code_2d):
    """Human-readable sector name for a 2-digit NAICS code."""
    return SECTOR_NAMES.get(code_2d, f"Sector {code_2d}")


# ---------------------------------------------------------------------------
# Text processing
# ---------------------------------------------------------------------------

def stem(word):
    """Minimal suffix stripping for NAICS vocabulary matching."""
    for suffix in ("turing", "ation", "ment", "ting", "ning", "ring",
                   "sion", "ness", "iers", "ies", "ers", "ing", "ion",
                   "als", "ous", "ble", "ive", "ful",
                   "ed", "ly", "es", "er", "al"):
        if word.endswith(suffix) and len(word) - len(suffix) >= 3:
            return word[:-len(suffix)]
    if word.endswith("s") and len(word) > 4:
        return word[:-1]
    return word


def tokenize(text):
    """Extract meaningful tokens from text, with basic stemming and synonym expansion."""
    words = set(re.findall(r"[a-z]{3,}", text.lower()))
    words = words - ANALYSIS_STOPWORDS
    stemmed = set()
    for w in words:
        stemmed.add(w)
        s = stem(w)
        if s != w:
            stemmed.add(s)
        if w in SYNONYM_BRIDGE:
            stemmed.update(SYNONYM_BRIDGE[w])
    return stemmed


# ---------------------------------------------------------------------------
# NAICS Description Tree
# ---------------------------------------------------------------------------

def load_naics_descriptions(file_2017=None, file_2022=None):
    """Load US NAICS descriptions from official XLSX files.

    Uses 2017 as primary, 2022 fills gaps.
    Expands combined sector codes (31-33, 44-45, 48-49) into individual codes.
    """
    import openpyxl
    if file_2017 is None:
        file_2017 = NAICS_2017_FILE
    if file_2022 is None:
        file_2022 = NAICS_2022_FILE

    descs = {}
    for xlsx_file in [file_2017, file_2022]:
        if not os.path.exists(xlsx_file):
            continue
        wb = openpyxl.load_workbook(xlsx_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            raw_code = str(row[1]).strip() if row[1] else ""
            name = str(row[2]).strip() if row[2] else ""
            if not raw_code or not name:
                continue
            if "-" in raw_code:
                parts = raw_code.split("-")
                if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                    for code_num in range(int(parts[0]), int(parts[1]) + 1):
                        if str(code_num) not in descs:
                            descs[str(code_num)] = name
                continue
            if raw_code not in descs:
                descs[raw_code] = name
        wb.close()
    return descs


def build_desc_tree(naics_code, naics_descs):
    """Build description tree: collect names at 6, 5, 4, 3, 2 digit levels."""
    code = str(naics_code).strip()
    parts = []
    for length in (6, 5, 4, 3, 2):
        prefix = code[:length]
        if prefix in naics_descs:
            parts.append(naics_descs[prefix])
    return " ".join(parts)


# ---------------------------------------------------------------------------
# BLS DART Rate Loading
# ---------------------------------------------------------------------------

def load_dart_rates(filepath=None):
    """Load BLS SOII Table 1 DART rates from XLSX.

    Returns dict: NAICS code string -> DART rate (float).
    """
    if filepath is None:
        filepath = BLS_DART_FILE
    if not os.path.exists(filepath):
        return {}
    import openpyxl
    rates = {}
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_code = str(row[1]).strip().replace("\xa0", "") if row[1] else ""
        raw_rate = str(row[3]).strip() if row[3] else ""
        if not raw_code or not raw_rate or raw_rate == "-":
            continue
        try:
            rate_val = float(raw_rate)
        except ValueError:
            continue
        if "-" in raw_code:
            parts = raw_code.split("-")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                for code_num in range(int(parts[0]), int(parts[1]) + 1):
                    rates[str(code_num)] = rate_val
                continue
        if not raw_code.isdigit():
            continue
        rates[raw_code] = rate_val
    wb.close()
    return rates


def lookup_dart(naics_code, dart_rates):
    """Cascading DART rate lookup: 6 -> 5 -> 4 -> 3 -> 2 digit.

    Returns (rate, matched_prefix) or (None, None).
    """
    code = str(naics_code).strip()
    for length in (6, 5, 4, 3, 2):
        prefix = code[:length]
        if prefix in dart_rates:
            return dart_rates[prefix], prefix
    return None, None
